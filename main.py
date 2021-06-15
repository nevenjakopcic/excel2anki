import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox

import genanki
from openpyxl import load_workbook


def upload():
    path = filedialog.askopenfilename()
    global ws, filename
    if path.endswith('.xlsx'):
        ws = load_workbook(path).active
        filename = get_file_name(path)
    else:
        messagebox.showerror('Error', 'Please select an .xlsx file.')


def get_file_name(path):
    return os.path.basename(path).split('.')[0]


def generate():
    global ws, filename
    deck = genanki.Deck(int(time.time()), 'Test deck')
    if ws:
        row = 1
        while ws.cell(row=row, column=1).value:
            deck.add_note(genanki.Note(
                model=anki_model,
                fields=[
                    ws.cell(row=row, column=1).value,
                    ws.cell(row=row, column=2).value]))
            row = row + 1

        genanki.Package(deck).write_to_file(os.path.join(os.path.expanduser("~/Desktop"), f'{filename}.apkg'))
        messagebox.showinfo('Success', 'Successfully generated file.')
    else:
        messagebox.showerror('Error', 'Please select a file.')


anki_model = genanki.Model(
    int(time.time()),
    'Basic Model',
    fields=[
        {'name': 'Question'},
        {'name': 'Answer'}],
    templates=[{
        'name': 'Card',
        'qfmt': '{{Question}}',
        'afmt': '{{FrontSide}}<hr id="answer">{{Answer}}'}])
ws = None
filename = 'output'
root = tk.Tk()
tk.Button(root, text='Open', command=upload).grid(row=2, column=0, sticky=tk.W, pady=4)
tk.Button(root, text='Generate', command=generate).grid(row=2, column=1, sticky=tk.W, pady=4)

root.mainloop()
